import openpyxl
import os
import tkinter as tk
from win32com import client
from openpyxl.styles import Alignment
from tkinter import ttk
import ctypes



root = tk.Tk()
root.title("Generador de Cotizaciones")

p1 = tk.PhotoImage(file = "./altiuslogo.png")
root.iconphoto(False, p1)

termsLabels = []
entries = []
entriesCliente = []
entriesProductos = []
checkBoxes = []
entriesTerms = []
dropDowns = []
dropDownsProducts = []
offset = 160
entryWidth = 40
canvasWidth = 1250
frame = tk.Frame(root, width = canvasWidth, height = 800)
frame.pack(expand=True, fill=tk.BOTH)
canvas1 = tk.Canvas(frame, width=canvasWidth, height=800, relief="raised",scrollregion=(0,0,500,800))
vbar=tk.Scrollbar(frame,orient=tk.VERTICAL)
vbar.pack(side=tk.RIGHT,fill=tk.Y)
vbar.config(command=canvas1.yview)

canvas1.config(yscrollcommand=vbar.set)

totDict = {}
currencies = []

productValues ={
    }

allProducts = []


def firstScreen():
    # Textos e inputs

    canvas1.pack()

    titleLabel = tk.Label(root, text="Datos Generales")
    titleLabel.config(font=("helvetica", 14))
    canvas1.create_window(90, 20, window=titleLabel)

    names = ["Date", "Quote #", "Valid Until", "Tipo", "Origen", "Destino", "Equipo", "Peso", "T/T"]

    for i in range(0, len(names)):
        entry = tk.Entry(root, width=entryWidth)
        entries.append(entry)

    x = 50
    y = 50

    for i in range(0, len(names)):
        label = tk.Label(root, text=names[i])
        label.config(font=('helvetica', 10))
        canvas1.create_window(x, y, window=label)
        canvas1.create_window(x + offset, y, window=entries[i])
        y += 25

    buttonOffsetY = 80
    buttonOffsetX = 120


    button1 = tk.Button(text = "Generar documento", command = addToExcel, height=5, width=16)
    button1.config(font=('helvetica', 18))
    canvas1.create_window(640, 690, window=button1)
    canvas1.create_rectangle(639 + buttonOffsetX, 690 + buttonOffsetY, 640 - buttonOffsetX, 690 - buttonOffsetY, fill = "black")

def secondScreen():
    names = ["Nombre", "Calle", "Distrito", "Ciudad", "País"]

    x = 50
    y = 350

    titleLabel = tk.Label(root, text="Información del Cliente")
    titleLabel.config(font=("helvetica", 14))
    canvas1.create_window(110, y-50, window=titleLabel)



    for i in range(0, len(names)):
        entry = tk.Entry(root, width=entryWidth)
        entriesCliente.append(entry)

    for i in range(0, len(names)):
        label = tk.Label(root, text=names[i])
        label.config(font=('helvetica', 10))
        canvas1.create_window(x, y, window=label)
        canvas1.create_window(x + offset, y, window=entriesCliente[i])
        y += 25

def thirdScreen():
    y = 45
    createThirdScreenLabels()
    newProductEntry(45)

    buttonAdd = tk.Button(text="+", command= lambda: newProductEntry(y + ((len(entriesProductos)) * 25)))
    canvas1.create_window(1030, 70, window=buttonAdd)

    buttonRemove = tk.Button(text="-", command= removeEntry)
    canvas1.create_window(1080, 70, window=buttonRemove)


    #printButton = tk.Button(text="print", command= printProducts)
    #canvas1.create_window(1000, 500, window=printButton)

def createThirdScreenLabels():

    x = 500
    y = 45

    titleLabel = tk.Label(root, text="Conceptos")
    titleLabel.config(font=("helvetica", 14))
    canvas1.create_window(x, 20, window=titleLabel)

    descriptionLabel = tk.Label(root, text="Descripción")
    descriptionLabel.config(font=("helvetica", 11))
    canvas1.create_window(x, y, window=descriptionLabel)
    x += 250

    precioLabel = tk.Label(root, text="Precio Unit")
    precioLabel.config(font=("helvetica", 11))
    canvas1.create_window(x, y, window=precioLabel)
    x += 100

    cantidadLabel = tk.Label(root, text="Cantidad")
    cantidadLabel.config(font=("helvetica", 11))
    canvas1.create_window(x, y, window=cantidadLabel)
    x += 75

    igvLabel = tk.Label(root, text="IGV")
    igvLabel.config(font=("helvetica", 11))
    canvas1.create_window(x, y, window=igvLabel)
    x += 50

    monedaLabel = tk.Label(root, text="Moneda")
    monedaLabel.config(font=("helvetica", 11))
    canvas1.create_window(x, y, window=monedaLabel)

def changePrice(i, chosenValue):
    value = ""
    if(chosenValue.upper() in productValues):
        value = productValues[chosenValue.upper()]
    x = entriesProductos[i][1]
    x.delete(0, 'end')
    x.insert(tk.END, value)

def newProductEntry(y):

    currentPosition = len(entriesProductos)
    if(len(entriesProductos) < 16):
        productoEntry = []
        x = 500

        chosenProduct = tk.StringVar(root)
        chosenProduct.set(allProducts[0])
        setPrice = productValues[chosenProduct.get()]
        chosenProduct.trace('w', lambda var_name, var_index, operation: changePrice(currentPosition, chosenProduct.get()))
        dropdownProduct = ttk.Combobox(root, width=30,textvariable=chosenProduct)
        dropdownProduct["values"] = allProducts
        dropDownsProducts.append(dropdownProduct)
        productoEntry.append(chosenProduct)
        canvas1.create_window(x+60, y + 25, window=dropdownProduct)

        #entryDescription = tk.Entry(root, width=40)
        #productoEntry.append(entryDescription)
        #canvas1.create_window(x + 80, y + 25, window=entryDescription)
        x += 250

        entryPrice = tk.Entry(root, width=10)
        productoEntry.append(entryPrice)

        entryPrice.insert(tk.END, productValues[chosenProduct.get()])
        canvas1.create_window(x, y + 25, window=entryPrice)
        x += 100

        entryCantidad = tk.Entry(root, width=10)
        productoEntry.append(entryCantidad)
        canvas1.create_window(x, y + 25, window=entryCantidad)
        x += 75

        i = tk.IntVar()
        checkBoxIGV = tk.Checkbutton(root, variable = i)
        checkBoxes.append(checkBoxIGV)
        productoEntry.append(i)
        canvas1.create_window(x+2, y + 25, window=checkBoxIGV)
        x += 50

        chosenMoneda = tk.StringVar(root)
        chosenMoneda.set("USD")
        dropdownMoneda = tk.OptionMenu(root, chosenMoneda, "USD", "PEN", "EUR")
        dropDowns.append(dropdownMoneda)
        canvas1.create_window(x, y + 25, window=dropdownMoneda)
        #entryMoneda = tk.Entry(root, width=10)
        productoEntry.append(chosenMoneda)
        #canvas1.create_window(x, y + 25, window=entryMoneda)

        entriesProductos.append(productoEntry)
        print(entriesProductos)

def removeEntry():
    currentEntries = len(entriesProductos)
    if(currentEntries > 1):
        i = 0
        for x in entriesProductos[currentEntries-1]:
            if(i!=3 and i!= 4 and i!=0):
                x.destroy()
            elif(i == 3):
                checkBoxes[currentEntries - 1].destroy()
                checkBoxes.pop(currentEntries-1)
            elif(i == 4):
                dropDowns[currentEntries - 1].destroy()
                dropDowns.pop(currentEntries-1)
            elif(i == 0):
                dropDownsProducts[currentEntries - 1].destroy()
                dropDownsProducts.pop(currentEntries - 1)
            i+=1

        entriesProductos.pop(currentEntries-1)

def fourthScreen():

    titleLabel = tk.Label(root, text="Terminos y condiciones")
    titleLabel.config(font=("helvetica", 14))
    canvas1.create_window(130, 500, window=titleLabel)

    y = 500

    newTermEntry(y)

    buttonAdd = tk.Button(text="+", command=lambda: newTermEntry(y + ((len(entriesTerms)) * 25)))
    canvas1.create_window(540, 535, window=buttonAdd)

    buttonRemove = tk.Button(text="-", command=delTermEntry)
    canvas1.create_window(590, 535, window=buttonRemove)

def newTermEntry(y):
    x = 130

    if len(entriesTerms) < 10:
        entryTerms = tk.Entry(root, width=80)
        entriesTerms.append(entryTerms)
        numberLabel = tk.Label(root, text=len(entriesTerms))
        numberLabel.config(font=("helvetica", 10))
        termsLabels.append(numberLabel)
        canvas1.create_window(x + 140, y + 35, window=entryTerms)
        canvas1.create_window(x - 115, y + 35, window=numberLabel)

def delTermEntry():
            currentNum = len(entriesTerms)-1
            x = entriesTerms[currentNum]
            label = termsLabels[currentNum]
            entriesTerms.pop(currentNum)
            termsLabels.pop(currentNum)
            x.destroy()
            label.destroy()

def printProducts():
    for x in entriesProductos:
        for y in x:
                print(y.get(), end=" ")
        print("")
    return

def formatNumber(x):
    if(x == "0"):
        return "0"
    decimals = x.split(".", 1)
    counter = 1
    for i in range(len(decimals[0]) - 1, 0, -1):
        if (counter % 3 == 0):
            decimals[0] = decimals[0][:i] + "," + decimals[0][i:]
        counter += 1
    x = decimals[0] + "." + decimals[1]
    if (len(decimals[1]) < 2):
        x += "0"
    return x

def convertToPdf(filename):
    directory = os.getcwd()
    excel = client.Dispatch("Excel.Application")

    input_file = r'' + directory + filename + '.xlsx'
    output_file = r'' + directory + filename + '.pdf'

    print(output_file)
    print(filename)
    print(input_file)

    sheets = excel.Workbooks.Open(input_file)
    worksheet = sheets.Worksheets[0]

    try:
        worksheet.ExportAsFixedFormat(0, output_file)
    except:
        ctypes.windll.user32.MessageBoxW(0, "Cerrar PDF antes de generar otro", "Error", 1)

    sheets.Close(SaveChanges=False)
    excel.Quit()

def addToExcel():

    wb = openpyxl.load_workbook('plantilla.xlsx')
    wb2 = openpyxl.load_workbook('plantilla.xlsx')
    sheet = wb["Quote 1"]
    costosSheet = wb2["Costo"]

    #Add Datos Generales

    for i in range(2, 11):
        nombre = sheet.cell(row=i, column=6).value
        valor = entries[i-2].get()
        sheet.cell(row=i, column=7).value = valor

    for x in entries:
        print(x.get())
    filename = './quote' + entries[1].get()

    #Add Informacion de Cliente

    for i in range(7, 10):
        valor = entriesCliente[i-7].get()
        sheet.cell(row=i, column= 1).value = valor

    cell = sheet.cell(row=10, column = 1)
    if(entriesCliente[3].get() != "" and entriesCliente[4].get() != ""):
        cell.value = entriesCliente[3].get() + " - " + entriesCliente[4].get()
    elif(entriesCliente[3].get() != "" and entriesCliente[4].get() == ""):
        cell.value = entriesCliente[3].get()
    elif(entriesCliente[3].get() == "" and entriesCliente[4].get() != ""):
        cell.value = entriesCliente[4].get()
    else:
        cell.value = ""

    #Add Products

    curColumn = 1
    curRow = 14
    for x in entriesProductos:
        newProductValue = ''
        newProductName = ''
        for y in x:
            cell = sheet.cell(row=curRow, column=curColumn)
            valor = y.get()
            if (curColumn == 5):
                if (valor == 1):
                    valor = "X"
                else:
                    valor = ""
            if(valor != ""):
                if(curColumn == 3):
                    cell.alignment = Alignment(horizontal="right")
                    newProductValue = valor
                    valor = formatNumber(str(round(float(valor),2)))
                if(curColumn == 6):
                    totDict[valor] = 0
                    totDict[valor+"IGV"] = 0
                    if(valor not in currencies and valor != ""):
                        currencies.append(valor)
                if(curColumn == 1):
                    valor = valor.upper()
                    if(valor not in productValues):
                        newProductName = valor
                cell.value = valor
            curColumn += 1
            if(curColumn == 2):
                curColumn += 1

        if(newProductName != '' and newProductValue != ''):
            nameCell = costosSheet.cell(row=len(allProducts)+1, column=1)
            nameCell.value = newProductName
            valueCell = costosSheet.cell(row=len(allProducts)+1, column=2)
            valueCell.value = newProductValue
            productValues[newProductName] = newProductValue
            allProducts.append(newProductName)
        curColumn = 1
        curRow += 1

    #Add Monto
    curRow = 14
    for x in entriesProductos:
        if(x[1].get() != "" and x[2].get() != ""):
            monto = float(x[1].get()) * float(x[2].get())
            cell = sheet.cell(row=curRow, column=7)
            cell.alignment = Alignment(horizontal="right")
            cell.value = monto
            curRow+=1
            key = x[4].get()
            totDict[key] += monto
            if(x[3].get() == 1):
                monto *= 0.18
                key += "IGV"
                totDict[key] += monto

    print(totDict)


    #Add Totales
    curRow = 30
    curColumn = 5
    numCurrencies = len(currencies)-1
    for x in totDict:
        if("IGV" in x):
            totalCell = sheet.cell(row=curRow+numCurrencies, column=curColumn+2)
            cell = sheet.cell(row=curRow+numCurrencies, column=curColumn)
            totalCell.value = formatNumber(str(round(totDict[x],2)))
            cell.value = "IGV " + x.replace("IGV", "")

        else:
            totalCell = sheet.cell(row=curRow, column=curColumn + 2)
            totalCell.value = formatNumber(str(round(totDict[x],2)))
            cell = sheet.cell(row=curRow, column=curColumn)
            cell.value = "Subtotal " + x
            curRow += 1
        totalCell.alignment = Alignment(horizontal="right")

    #Add Total Final
    for i in range(0, len(currencies)):
        cell = sheet.cell(row=i+37, column=5)
        cell.value = "TOTAL " + str(currencies[i])
        cell = sheet.cell(row=i+37, column=7)
        cell.alignment = Alignment(horizontal="right")
        print(currencies[i])
        totalSum = (totDict[currencies[i]] + totDict[currencies[i]+"IGV"])
        cell.number_format = '#,##0.00'
        stringSum = str(round(totalSum,2))
        cell.value = formatNumber(stringSum)

    #Add Terminos
    curRow = 32
    curColumn = 1
    print(len(entriesTerms))
    for i in range(0, len(entriesTerms)):
        cell = sheet.cell(row=curRow, column=curColumn)
        cell.value = str(i+1) +". "+entriesTerms[i].get()
        curRow += 1



    wb.save(filename+".xlsx")
    wb2.save("./plantilla.xlsx")

    filename = "\quote"+entries[1].get()


    convertToPdf(filename)

def readProductValues():
    wb = openpyxl.load_workbook('plantilla.xlsx')
    costosSheet = wb["Costo"]

    curRow = 1
    cell = costosSheet.cell(1,1)
    while(cell.value != '' and cell.value != None):
        print(cell.value)
        valor = ""
        key = ""
        for curColumn in range(1,3):
            cell = costosSheet.cell(row=curRow, column=curColumn)
            if(curColumn == 1):
                key = cell.value
            else:
                valor = cell.value
        productValues[key] = valor
        curRow+=1
        cell = costosSheet.cell(row=curRow, column=curColumn)



readProductValues()

for x in productValues:
    allProducts.append(x)


firstScreen()

secondScreen()

thirdScreen()

fourthScreen()



root.mainloop()


